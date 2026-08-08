from __future__ import annotations
import csv, hashlib, io, json, unittest, zipfile
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
class TestLive10Evidence(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.pkg=json.loads((ROOT/'evidence/B5_SITE_ANALYZER_WAVE2_LIVE_10_EVIDENCE_PACKAGE_V1.json').read_text())
        cls.live_run=cls.pkg['run']; cls.preview=cls.pkg['preview']; cls.manifest=cls.pkg['export_manifest']
        cls.j=json.loads((ROOT/'exports/live_10_edited.json').read_text())
        cls.c=list(csv.DictReader(io.StringIO((ROOT/'exports/live_10_edited.csv').read_text(encoding='utf-8-sig'))))
        wb=load_workbook(ROOT/'exports/live_10_edited.xlsx',read_only=True,data_only=True); ws=wb['Live10']; cls.x=list(ws.iter_rows(min_row=2,values_only=True)); wb.close()
    def test_01_actual_loopback_http(self): self.assertTrue(self.manifest['b4_binding']['actual_loopback_http_run'])
    def test_02_b4_server_blob_bound(self): self.assertEqual(self.live_run['b4_server_blob'],'671f05d903dcd8aecd7fd8160b493981f34a5b5a')
    def test_03_live_total_10(self): self.assertEqual(self.live_run['record_count'],10)
    def test_04_retry_once(self): self.assertEqual(self.live_run['retry_count'],1)
    def test_05_http_status_contains_503_then_200(self): self.assertEqual(self.live_run['http_attempts'][1]['statuses'],[503,200])
    def test_06_preview_10(self): self.assertEqual(self.preview['rowCount'],10)
    def test_07_json_10(self): self.assertEqual(len(self.j),10)
    def test_08_csv_10(self): self.assertEqual(len(self.c),10)
    def test_09_xlsx_10(self): self.assertEqual(len(self.x),10)
    def test_10_order_parity(self): self.assertEqual([r['id'] for r in self.j],[int(r['id']) for r in self.c]); self.assertEqual([r['id'] for r in self.j],[r[0] for r in self.x])
    def test_11_edit_parity(self): self.assertEqual((self.preview['rows'][4]['cells'][2]['value'],self.j[4]['edited_value'],int(self.c[4]['edited_value']),self.x[4][2]),(5550,5550,5550,5550))
    def test_12_field_name_parity(self): self.assertEqual(self.preview['columns'][1]['label'],'item_name'); self.assertEqual(self.c[0]['item_name'],self.x[0][1])
    def test_13_source_row_pointers(self): self.assertEqual(sum(bool(r['__source']['rowPointer']) for r in self.j),10)
    def test_14_source_element_pointers(self): self.assertEqual(sum(len(r['__source']['elementPointers']) for r in self.j),40)
    def test_15_xlsx_openxml_and_sha(self):
        p=ROOT/'exports/live_10_edited.xlsx'; self.assertTrue(zipfile.is_zipfile(p)); self.assertEqual(hashlib.sha256(p.read_bytes()).hexdigest(),self.manifest['files']['xlsx']['binary_sha256'])
if __name__=='__main__': unittest.main()
